# app.py — Waybill Maker (multi-parser: Japafrica + ZF Scandi)
import io, re, statistics
from dataclasses import dataclass
from typing import List, Optional, Tuple, Type

import fitz  # PyMuPDF
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook

# ───────────────── UI ─────────────────
st.set_page_config(page_title="Waybill Maker", page_icon="📦", layout="wide")
st.title("📦 Waybill Maker")

with st.sidebar:
    st.header("Парсеры")
    st.caption("Выходные колонки: MPN | Replacem | Quantity | Totalsprice | Order reference")

# ───────────────── Regex (общие) ─────────────────
# 11 цифр (допускаем ведущую C → берём только цифры)
RE_MPN_11D = re.compile(r"\bC?(\d{11})\b")

# 81.36304-0019 и т.п. (допускаем ведущую C → берём без неё)
RE_MPN_DOT_OPT_C = re.compile(r"\bC?(\d{2}\.\d{5}-\d{3,4})\b")

# ZF формат: 0750.117.859 / 4475.305.212
RE_MPN_ZF = re.compile(r"\b\d{3,4}\.\d{3}\.\d{3}\b")

RE_INT   = re.compile(r"^\d{1,4}$")
RE_DEC   = re.compile(r"^\d{1,6}[.,]\d{2}$")
RE_MONEY = re.compile(r"\d{1,3}(?:[ \u00A0]?\d{3})*[.,]\d{2}")

# Заголовки LV/EN (для Fallback)
RE_HDR_ART_LV = re.compile(r"(?i)artik|artikul")
RE_HDR_QTY_LV = re.compile(r"(?i)daudz")
RE_HDR_SUM_LV = re.compile(r"(?i)summa|summ")

RE_HDR_PART_EN = re.compile(r"(?i)\bpart\b|ref\.*\s*ª?\s*pe[cç]a")
RE_HDR_QTY_EN  = re.compile(r"(?i)\bqty\b|quant")
RE_HDR_SUM_EN  = re.compile(r"(?i)\beur\b|\bamount\b|\btotal\b|\bsum\b")

# Order reference (например: ORDER 126152 / #126152 / просто 1******)
RE_ORDER = [
    re.compile(r"(?:^|\s)#\s*(1\d{5})(?:\s|$)"),
    re.compile(r"(?i)\border[_\-\s]*0*(1\d{5})"),
    re.compile(r"(?<![\d.,])(1\d{5})(?![\d.,])"),
]

# ───────────────── Utils ─────────────────
def to_float(s: str) -> float:
    return float(s.replace("\u00A0"," ").replace(" ","").replace(".","").replace(",","."))

def to_int(s: str) -> int:
    return int(round(to_float(s)))

def fmt_money_dot(s: Optional[str]) -> str:
    if not s:
        return "0.00"
    return f"{to_float(s):.2f}"

@dataclass
class Word:
    x0: float; y0: float; x1: float; y1: float; text: str
@dataclass
class Line:
    y: float; words: List[Word]; text: str
@dataclass
class Band:
    name: str; x_left: float; x_right: float
@dataclass
class OrderMark:
    x: float; y: float; value: str

# ───────────────── low-level text ─────────────────
def load_words_per_page(pdf_bytes: bytes) -> List[List[Word]]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    out=[]
    for p in doc:
        ws=[Word(w[0],w[1],w[2],w[3],w[4]) for w in p.get_text("words")]
        ws.sort(key=lambda w:(round(w.y0,1), w.x0))
        out.append(ws)
    return out

def group_lines(words: List[Word]) -> List[Line]:
    if not words: return []
    heights=[w.y1-w.y0 for w in words if (w.y1-w.y0)>0.2]
    h = statistics.median(heights) if heights else 8.0
    ytol=max(1.2, h*0.65)

    res=[]; cur=[]; last=None
    for w in words:
        if last is None or abs(w.y0-last)<=ytol:
            cur.append(w); last = w.y0 if last is None else (last+w.y0)/2
        else:
            cur.sort(key=lambda t:t.x0); res.append(cur); cur=[w]; last=w.y0
    if cur: cur.sort(key=lambda t:t.x0); res.append(cur)

    out=[]
    for ln in res:
        y=statistics.fmean([w.y0 for w in ln])
        out.append(Line(y=y, words=ln, text=" ".join(w.text for w in ln)))
    out.sort(key=lambda L:L.y)
    return out

def in_band(w: Word, b: Band) -> bool:
    cx=(w.x0+w.x1)/2
    return b.x_left<=cx<=b.x_right

def join_money_tokens(tokens: List[Word], gap_limit: int = 8) -> Optional[str]:
    if not tokens: return None
    tokens.sort(key=lambda w: w.x0)
    groups=[]; cur=[tokens[0]]
    for w in tokens[1:]:
        if (w.x0 - cur[-1].x1) <= gap_limit:
            cur.append(w)
        else:
            groups.append(cur); cur=[w]
    groups.append(cur)
    g = max(groups, key=lambda G: max(w.x1 for w in G))
    raw = "".join(w.text.replace("\u00A0","").replace(" ","") for w in g)
    if not re.search(r"[.,]\d{2}$", raw):
        if re.fullmatch(r"\d{1,9}", raw):
            raw = raw + ".00"
        else:
            raw = re.sub(r"(\d{2})$", r".\1", raw)
    return raw

def pick_total_for_line(line: Line, sum_band: Band) -> Optional[str]:
    cands=[w for w in line.words if in_band(w,sum_band) and (RE_MONEY.fullmatch(w.text) or RE_INT.fullmatch(w.text) or RE_DEC.fullmatch(w.text))]
    if not cands: return None
    tok=join_money_tokens(cands, gap_limit=8)
    if not tok: return None
    # «1» + «027,07» → «1027,07»
    lefts=[w for w in line.words if w.x1<=cands[-1].x0+1 and (cands[-1].x0-w.x1)<=8]
    if lefts:
        lefts.sort(key=lambda w:w.x1, reverse=True)
        Lw=lefts[0]
        if re.fullmatch(r"[1-9]", Lw.text) and re.fullmatch(r"0\d{2}[.,]\d{2}", tok):
            tok=Lw.text+tok
    return tok

def order_from_text(txt: str) -> Optional[str]:
    for p in RE_ORDER:
        m=p.search(txt)
        if m: return m.group(1)
    return None

def collect_order_marks(lines: List[Line]) -> List[OrderMark]:
    out=[]
    for L in lines:
        val=order_from_text(L.text)
        if val:
            xs=[(w.x0+w.x1)/2 for w in L.words if any(p.search(w.text) for p in RE_ORDER)]
            cx=statistics.median(xs) if xs else statistics.fmean([(w.x0+w.x1)/2 for w in L.words])
            out.append(OrderMark(cx, L.y, val))
    if not out: return []
    x_med=statistics.median([m.x for m in out])
    col=[m for m in out if abs(m.x-x_med)<=42]
    out = col if len(col)>=max(3, len(out)//2) else out
    out.sort(key=lambda m:m.y)
    return out

def nearest_order_above(marks: List[OrderMark], y: float) -> str:
    """
    Ближайший order сверху относительно y (или ближайший по вертикали, но не дальше 30 pt).
    """
    prev=[m for m in marks if m.y <= y + 2]
    if prev:
        return prev[-1].value
    if not marks:
        return ""
    best=min(marks, key=lambda m: abs(m.y - y))
    return best.value if abs(best.y - y) <= 30 else ""

# ───────────────── БАЗА ДЛЯ ПАРСЕРОВ ─────────────────
class BaseParser:
    NAME = "Base"

    def matches(self, lines: List[Line], words: List[Word]) -> bool:
        return False

    def parse_page(self, lines: List[Line], words: List[Word]) -> List[dict]:
        raise NotImplementedError

    # Общая логика для извлечения строк по коридорам колонок
    def _extract_rows_by_bands(self, lines: List[Line], bands: List[Band]) -> List[dict]:
        B = {b.name:b for b in bands}
        orders = collect_order_marks(lines)
        rows=[]

        def find_mpn(line_text: str):
            m = RE_MPN_11D.search(line_text) \
                or RE_MPN_DOT_OPT_C.search(line_text) \
                or RE_MPN_ZF.search(line_text)
            if not m:
                return None
            return m.group(1) if m.lastindex else m.group(0)

        cand_idx=[i for i,L in enumerate(lines) if find_mpn(L.text)]

        for i in cand_idx:
            L=lines[i]
            mpn = find_mpn(L.text)
            if not mpn:
                continue

            # Quantity
            best=(1e9,None)
            for d in [0,1]:
                for sgn in (0,-1,1):
                    j=i+sgn*d
                    if j<0 or j>=len(lines): continue
                    for w in lines[j].words:
                        if in_band(w,B["Daudz."]) and (RE_INT.fullmatch(w.text) or RE_DEC.fullmatch(w.text)):
                            dy=abs(lines[j].y - L.y)
                            if dy<best[0]:
                                best=(dy,w.text)
                if best[1]: break
            qty = to_int(best[1]) if best[1] else None

            # Totalsprice
            bestT=(1e9,None)
            for d in [0,1]:
                for sgn in (0,-1,1):
                    j=i+sgn*d
                    if j<0 or j>=len(lines): continue
                    tok=pick_total_for_line(lines[j], B["Summa"])
                    if tok:
                        dy=abs(lines[j].y - L.y)
                        if dy<bestT[0]:
                            bestT=(dy,tok)
                if bestT[1]: break
            total = fmt_money_dot(bestT[1]) if bestT[1] else None

            if qty is None or total is None:
                continue

            order = nearest_order_above(orders, L.y)

            # если total численно совпал с qty — попробуем правую склейку
            try:
                if abs(to_int(bestT[1]) - qty) == 0:
                    c=[w for w in L.words if in_band(w,B["Summa"]) and (RE_MONEY.fullmatch(w.text) or RE_INT.fullmatch(w.text) or RE_DEC.fullmatch(w.text))]
                    if len(c)>=2:
                        total = fmt_money_dot(join_money_tokens(c))
            except:
                pass

            rows.append({
                "MPN": mpn,
                "Replacem": "",
                "Quantity": qty,
                "Totalsprice": total,
                "Order reference": order
            })
        return rows

# ───────────────── КОНКРЕТНЫЕ ПАРСЕРЫ ─────────────────
# 1) ZF Scandi / ZF Danmark (Sales Invoice)
class ZFScandiParser(BaseParser):
    NAME = "ZF Scandi (ZF Danmark)"

    def matches(self, lines: List[Line], words: List[Word]) -> bool:
        head = "\n".join(L.text for L in lines[:80])
        return ("Sales Invoice" in head and "ZF DANMARK" in head) or ("No. Description" in head and "Amount" in head)

    def detect_bands(self, lines: List[Line], words: List[Word]) -> Optional[List[Band]]:
        RE_DESC = re.compile(r"(?i)\bdescription\b")
        RE_QTY  = re.compile(r"(?i)\bqty\b|\bshipment\s*qty\b|\bquantity\b")
        RE_AMT  = re.compile(r"(?i)\bamount\b")

        def _centers_for(line: Line, pat: re.Pattern) -> Optional[float]:
            xs=[(w.x0+w.x1)/2 for w in line.words if pat.search(w.text)]
            return sum(xs)/len(xs) if xs else None

        for L in lines[:160]:
            if RE_DESC.search(L.text) and RE_QTY.search(L.text) and RE_AMT.search(L.text):
                cx_d=_centers_for(L, RE_DESC); cx_q=_centers_for(L, RE_QTY); cx_a=_centers_for(L, RE_AMT)
                centers=[(n,c) for n,c in [("Desc",cx_d),("Qty",cx_q),("Amt",cx_a)] if c is not None]
                if len(centers) >= 2:
                    centers.sort(key=lambda t:t[1])
                    bands=[]
                    for i,(n,cx) in enumerate(centers):
                        left  = (centers[i-1][1]+cx)/2 if i>0 else cx-120
                        right = (cx+centers[i+1][1])/2 if i<len(centers)-1 else cx+200
                        bands.append(Band(n,left,right))
                    bands_sorted = sorted(bands, key=lambda b: b.x_left)
                    desc_band = next(b for b in bands_sorted if b.name=="Desc")
                    qty_band  = next(b for b in bands_sorted if b.name=="Qty")
                    amt_band  = next(b for b in bands_sorted if b.name=="Amt")

                    x_min = min(w.x0 for w in words)
                    artic_band = Band("Artikuls", x_min-10, (desc_band.x_left + qty_band.x_left)/2)

                    return [
                        artic_band,
                        Band("Daudz.", qty_band.x_left, qty_band.x_right),
                        Band("Summa",  amt_band.x_left, amt_band.x_right),
                    ]
        return None

    def parse_page(self, lines: List[Line], words: List[Word]) -> List[dict]:
        bands = self.detect_bands(lines, words)
        if not bands:
            return []
        return self._extract_rows_by_bands(lines, bands)

# 2) Japafrica (FACTURA / INVOICE)
class JapafricaParser(BaseParser):
    NAME = "Japafrica"

    def matches(self, lines: List[Line], words: List[Word]) -> bool:
        head = "\n".join(L.text for L in lines[:100])
        return ("JAPAFRICA" in head) and (("FACTURA" in head) or ("INVOICE" in head)) and ("QTY" in head and "EUR" in head)

    def detect_bands(self, lines: List[Line], words: List[Word]) -> Optional[List[Band]]:
        RE_PART = re.compile(r"(?i)ref\.*\s*ª?\s*pe[cç]a|part\s*no")
        RE_QTY  = re.compile(r"(?i)\bqty\b|quant")
        RE_EUR  = re.compile(r"(?i)\beur\b|\bamount\b|\btotal\b")

        def _centers_for(line: Line, pat: re.Pattern) -> Optional[float]:
            xs=[(w.x0+w.x1)/2 for w in line.words if pat.search(w.text)]
            return sum(xs)/len(xs) if xs else None

        for L in lines[:180]:
            if RE_PART.search(L.text) and RE_QTY.search(L.text) and RE_EUR.search(L.text):
                cx_p=_centers_for(L,RE_PART); cx_q=_centers_for(L,RE_QTY); cx_e=_centers_for(L,RE_EUR)
                centers=[(n,c) for n,c in [("Part",cx_p),("Qty",cx_q),("Eur",cx_e)] if c is not None]
                if len(centers) >= 2:
                    centers.sort(key=lambda t:t[1])
                    bands=[]
                    for i,(n,cx) in enumerate(centers):
                        left  = (centers[i-1][1]+cx)/2 if i>0 else cx-100
                        right = (cx+centers[i+1][1])/2 if i<len(centers)-1 else cx+200
                        bands.append(Band(n,left,right))
                    bands_sorted = sorted(bands, key=lambda b: b.x_left)
                    part_band = next(b for b in bands_sorted if b.name=="Part")
                    qty_band  = next(b for b in bands_sorted if b.name=="Qty")
                    eur_band  = next(b for b in bands_sorted if b.name=="Eur")
                    return [
                        Band("Artikuls", part_band.x_left, part_band.x_right),  # C81.36400-6007 → 81.36400-6007
                        Band("Daudz.",   qty_band.x_left,  qty_band.x_right),
                        Band("Summa",    eur_band.x_left,  eur_band.x_right),
                    ]
        return None

    def parse_page(self, lines: List[Line], words: List[Word]) -> List[dict]:
        bands = self.detect_bands(lines, words)
        if not bands:
            return []
        rows = self._extract_rows_by_bands(lines, bands)
        # убрать ведущую 'C' у MPN вида C81.36400-6007
        for r in rows:
            if r["MPN"] and isinstance(r["MPN"], str) and r["MPN"].startswith("C") and re.fullmatch(r"C\d{2}\.\d{5}-\d{3,4}", r["MPN"]):
                r["MPN"] = r["MPN"][1:]
        return rows

# 3) Fallback — универсальный LV/EN
class FallbackParser(BaseParser):
    NAME = "Fallback (generic LV/EN)"

    def _centers_for(self, line: Line, pat: re.Pattern) -> Optional[float]:
        xs=[(w.x0+w.x1)/2 for w in line.words if pat.search(w.text)]
        return sum(xs)/len(xs) if xs else None

    def detect_bands_lv(self, lines: List[Line], words: List[Word]) -> Optional[List[Band]]:
        for L in lines[:120]:
            if RE_HDR_ART_LV.search(L.text) and RE_HDR_QTY_LV.search(L.text) and RE_HDR_SUM_LV.search(L.text):
                cx_a=self._centers_for(L,RE_HDR_ART_LV); cx_q=self._centers_for(L,RE_HDR_QTY_LV); cx_s=self._centers_for(L,RE_HDR_SUM_LV)
                centers=[(n,c) for n,c in [("Artikuls",cx_a),("Daudz.",cx_q),("Summa",cx_s)] if c is not None]
                if len(centers)>=2:
                    centers.sort(key=lambda t:t[1])
                    bands=[]
                    for i,(n,cx) in enumerate(centers):
                        left  = (centers[i-1][1]+cx)/2 if i>0 else cx-90
                        right = (cx+centers[i+1][1])/2 if i<len(centers)-1 else cx+180
                        bands.append(Band(n,left,right))
                    for b,nm in zip(sorted(bands,key=lambda b:b.x_left),["Artikuls","Daudz.","Summa"]):
                        b.name=nm
                    return bands
        return None

    def detect_bands_en(self, lines: List[Line], words: List[Word]) -> Optional[List[Band]]:
        for L in lines[:140]:
            if RE_HDR_PART_EN.search(L.text) and RE_HDR_QTY_EN.search(L.text) and RE_HDR_SUM_EN.search(L.text):
                cx_p=self._centers_for(L,RE_HDR_PART_EN); cx_q=self._centers_for(L,RE_HDR_QTY_EN); cx_s=self._centers_for(L,RE_HDR_SUM_EN)
                centers=[(n,c) for n,c in [("Part",cx_p),("Qty",cx_q),("Sum",cx_s)] if c is not None]
                if len(centers)>=2:
                    centers.sort(key=lambda t:t[1])
                    bands=[]
                    for i,(n,cx) in enumerate(centers):
                        left  = (centers[i-1][1]+cx)/2 if i>0 else cx-90
                        right = (cx+centers[i+1][1])/2 if i<len(centers)-1 else cx+180
                        bands.append(Band(n,left,right))
                else:
                    continue
                mapped=[]
                for b in sorted(bands, key=lambda b:b.x_left):
                    nm = "Artikuls" if b.name=="Part" else ("Daudz." if b.name=="Qty" else "Summa")
                    mapped.append(Band(nm,b.x_left,b.x_right))
                return mapped
        return None

    def fallback_bands(self, words: List[Word]) -> List[Band]:
        x_min=min(w.x0 for w in words); x_max=max(w.x1 for w in words); W=x_max-x_min
        return [
            Band("Artikuls", x_min-10, x_min+0.47*W),
            Band("Daudz.",   x_min+0.47*W, x_min+0.66*W),
            Band("Summa",    x_min+0.66*W, x_max+20),
        ]

    def matches(self, lines: List[Line], words: List[Word]) -> bool:
        return True  # всегда как запасной

    def parse_page(self, lines: List[Line], words: List[Word]) -> List[dict]:
        bands = self.detect_bands_lv(lines, words) or self.detect_bands_en(lines, words)
        if not bands:
            bands = self.fallback_bands(words)
        return self._extract_rows_by_bands(lines, bands)

# ───────────────── Реестр и роутер ─────────────────
PARSERS: List[Type[BaseParser]] = [
    ZFScandiParser,   # специфичный
    JapafricaParser,  # специфичный
    FallbackParser,   # общий
]

def parse_pdf_with_registry(pdf_bytes: bytes) -> pd.DataFrame:
    pages = load_words_per_page(pdf_bytes)
    all_rows=[]
    for words in pages:
        if not words: 
            continue
        lines = group_lines(words)

        # выбираем первый подходящий парсер
        parser = None
        for cls in PARSERS:
            inst = cls()
            if inst.matches(lines, words):
                parser = inst
                break
        if parser is None:
            parser = FallbackParser()

        all_rows.extend(parser.parse_page(lines, words))

    df = pd.DataFrame(all_rows).drop_duplicates(keep="last")
    if df.empty:
        df = pd.DataFrame(columns=["MPN","Replacem","Quantity","Totalsprice","Order reference"])
    else:
        df = df[["MPN","Replacem","Quantity","Totalsprice","Order reference"]]
        df.reset_index(drop=True, inplace=True)
    return df

# ───────────────── Streamlit App ─────────────────
pdf_file = st.file_uploader("Загрузить PDF-счёт", type=["pdf"])
tpl_file = st.file_uploader("Шаблон Excel (необязательно)", type=["xlsx"])

if pdf_file:
    pdf_bytes = pdf_file.read()  # читаем один раз
    try:
        df = parse_pdf_with_registry(pdf_bytes)
    except Exception as e:
        st.error(f"Не удалось разобрать PDF: {e}")
        df = pd.DataFrame(columns=["MPN","Replacem","Quantity","Totalsprice","Order reference"])

    st.subheader("Предпросмотр")
    st.dataframe(df, use_container_width=True)

    # CSV
    if not df.empty:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Скачать waybill.csv", data=csv_bytes, file_name="waybill.csv", mime="text/csv")

    # Excel (с шаблоном или без)
    if st.button("⬇️ Скачать Excel"):
        if tpl_file:
            wb=load_workbook(tpl_file); ws=wb.active
            if ws.max_row <= 1:
                ws.append(["MPN","Replacem","Quantity","Totalsprice","Order reference"])
        else:
            wb=Workbook(); ws=wb.active
            ws.append(["MPN","Replacem","Quantity","Totalsprice","Order reference"])

        for r in df.itertuples(index=False):
            ws.append(list(r))

        bio=io.BytesIO(); wb.save(bio)
        st.download_button("Скачать waybill.xlsx",
                           data=bio.getvalue(),
                           file_name="waybill.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("Загрузите счёт — поддержаны Japafrica и ZF Scandi; остальные форматы будут разобраны Fallback-парсером.")
